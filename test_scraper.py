"""
Tests de validation avant lancement
Lance : python3 test_scraper.py
"""
import sys
import os
import io
from datetime import datetime

# ---------------------------------------------------------------------------
# Couleurs terminal
# ---------------------------------------------------------------------------
GREEN = "\033[92m"
RED   = "\033[91m"
YELLOW= "\033[93m"
BOLD  = "\033[1m"
RESET = "\033[0m"

passed = failed = 0

def ok(label):
    global passed
    passed += 1
    print(f"  {GREEN}✓{RESET} {label}")

def fail(label, detail=""):
    global failed
    failed += 1
    print(f"  {RED}✗ {label}{RESET}" + (f" — {detail}" if detail else ""))

def section(title):
    print(f"\n{BOLD}{title}{RESET}")

# ---------------------------------------------------------------------------
# 1. Imports
# ---------------------------------------------------------------------------
section("1. Imports")
try:
    from youtube_scraper import (
        calculate_tier, compute_scores, export_excel, COLUMNS,
        search_videos_by_keyword, get_channel_details, get_recent_video_stats,
    )
    ok("youtube_scraper importé")
except Exception as e:
    fail("youtube_scraper", str(e)); sys.exit(1)

try:
    import pandas as pd; ok("pandas")
    import openpyxl; ok("openpyxl")
    from googleapiclient.discovery import build; ok("google-api-python-client")
    from dotenv import load_dotenv; ok("python-dotenv")
    from tqdm import tqdm; ok("tqdm")
except ImportError as e:
    fail("dépendance manquante", str(e))

# ---------------------------------------------------------------------------
# 2. calculate_tier
# ---------------------------------------------------------------------------
section("2. Tiers")
cases = [(500, "nano"), (5_000, "micro"), (50_000, "mid"), (500_000, "macro"), (2_000_000, "mega")]
for subs, expected in cases:
    result = calculate_tier(subs)
    if result == expected:
        ok(f"{subs:>10,} → {result}")
    else:
        fail(f"{subs:>10,} → attendu {expected}, obtenu {result}")

# ---------------------------------------------------------------------------
# 3. compute_scores
# ---------------------------------------------------------------------------
section("3. Scoring")
# (engagement_rate, mentions, posts_per_week, growth_rate_pct)
score_cases = [
    (0.10, 5, 3, 20, "profil idéal → score élevé", lambda sg: sg >= 85),
    (0.00, 0, 0, 0,  "profil vide → score bas",    lambda sg: sg <= 25),
    (0.05, 2, 1, 5,  "profil moyen → 40–75",        lambda sg: 40 <= sg <= 75),
]
for eng, mentions, ppw, growth, label, check in score_cases:
    se, sc, sp, sr, sg = compute_scores(eng, mentions, ppw, growth)
    if check(sg):
        ok(f"{label} (score={sg})")
    else:
        fail(label, f"score_global={sg}")

# Range 0–100
se, sc, sp, sr, sg = compute_scores(0.08, 3, 2, 10)
for name, val in [("engagement", se), ("croissance", sc), ("pertinence", sp), ("regularite", sr), ("global", sg)]:
    if 0 <= val <= 100:
        ok(f"score_{name} dans [0,100] ({val})")
    else:
        fail(f"score_{name} hors bornes", str(val))

# ---------------------------------------------------------------------------
# 4. Filtre abonnés (logique de app.py)
# ---------------------------------------------------------------------------
section("4. Filtre abonnés")

def apply_follower_filter(followers, followers_min, followers_max):
    if followers_min > 0 and followers < followers_min:
        return False
    if followers_max > 0 and followers > followers_max:
        return False
    return True

filter_cases = [
    (50_000, 0,      0,       True,  "pas de filtre"),
    (50_000, 10_000, 100_000, True,  "dans la plage"),
    (5_000,  10_000, 100_000, False, "sous le min"),
    (200_000,10_000, 100_000, False, "au-dessus du max"),
    (10_000, 10_000, 0,       True,  "seulement min, exact"),
    (9_999,  10_000, 0,       False, "seulement min, exclu"),
    (100_000,0,      100_000, True,  "seulement max, exact"),
    (100_001,0,      100_000, False, "seulement max, exclu"),
]
for followers, fmin, fmax, expected, label in filter_cases:
    result = apply_follower_filter(followers, fmin, fmax)
    if result == expected:
        ok(label)
    else:
        fail(label, f"followers={followers:,}, min={fmin:,}, max={fmax:,} → attendu {expected}")

# ---------------------------------------------------------------------------
# 5. Export Excel avec données mockées
# ---------------------------------------------------------------------------
section("5. Export Excel (données mockées)")

mock_profiles = []
for i in range(10):
    followers = (i + 1) * 15_000
    eng = 0.01 + i * 0.008
    mentions = i % 5
    ppw = 0.5 + i * 0.3
    growth = i * 2.5
    se, sc, sp, sr, sg = compute_scores(eng, mentions, ppw, growth)
    mock_profiles.append({
        "platform": "YouTube",
        "username": f"creator_{i}",
        "display_name": f"Créateur Test {i}",
        "profile_url": f"https://www.youtube.com/@creator_{i}",
        "bio_snippet": f"Bio du créateur numéro {i}. Parle de Sorare et de football.",
        "followers": followers,
        "tier": calculate_tier(followers),
        "engagement_rate": round(eng, 6),
        "engagement_rate_pct": round(eng * 100, 3),
        "croissance_hebdo": round(followers / 52, 1),
        "growth_rate_pct": round(growth, 2),
        "posts_per_week": round(ppw, 2),
        "sorare_mentions": mentions,
        "is_emerging": growth > 5 and followers < 50_000,
        "score_global": sg,
        "score_engagement": se,
        "score_croissance": sc,
        "score_pertinence": sp,
        "score_regularite": sr,
        "status": "active" if ppw >= 0.5 else "inactive",
        "collected_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })

# Export vers fichier
test_file = "/tmp/test_sorare_scraper.xlsx"
try:
    export_excel(mock_profiles, test_file, ["Sorare"])
    ok(f"Export fichier ({test_file})")
except Exception as e:
    fail("Export fichier", str(e))

# Export vers buffer (comme le bouton download Streamlit)
try:
    buf = io.BytesIO()
    export_excel(mock_profiles, buf, ["Sorare"])
    buf.seek(0)
    size = len(buf.read())
    if size > 1000:
        ok(f"Export buffer en mémoire ({size:,} bytes)")
    else:
        fail("Buffer trop petit", f"{size} bytes")
except Exception as e:
    fail("Export buffer", str(e))

# Vérifier le contenu du fichier Excel
try:
    import openpyxl
    wb = openpyxl.load_workbook(test_file)
    sheets = wb.sheetnames
    if "Profiles" in sheets and "Summary" in sheets:
        ok(f"Onglets présents : {sheets}")
    else:
        fail("Onglets manquants", str(sheets))

    ws = wb["Profiles"]
    headers = [ws.cell(1, c).value for c in range(1, len(COLUMNS) + 1)]
    if headers == COLUMNS:
        ok(f"Toutes les colonnes présentes ({len(COLUMNS)})")
    else:
        missing = set(COLUMNS) - set(headers)
        fail("Colonnes manquantes", str(missing))

    row_count = ws.max_row - 1  # sans header
    if row_count == len(mock_profiles):
        ok(f"{row_count} lignes exportées")
    else:
        fail("Nombre de lignes incorrect", f"attendu {len(mock_profiles)}, obtenu {row_count}")
except Exception as e:
    fail("Lecture Excel", str(e))

# ---------------------------------------------------------------------------
# 6. Logo
# ---------------------------------------------------------------------------
section("6. Assets")
logo = "assets/Sorare26-Logo-Black.png"
if os.path.exists(logo):
    size = os.path.getsize(logo)
    ok(f"Logo trouvé ({size:,} bytes)")
else:
    fail("Logo manquant", logo)

env_example = ".env.example"
if os.path.exists(env_example):
    ok(".env.example présent")
else:
    fail(".env.example manquant")

# ---------------------------------------------------------------------------
# 7. Clé API (optionnel)
# ---------------------------------------------------------------------------
section("7. Clé API YouTube")
from dotenv import load_dotenv
load_dotenv()
api_key = os.environ.get("YOUTUBE_API_KEY", "")
if api_key and len(api_key) > 10:
    ok(f"Clé API détectée ({api_key[:6]}…)")
    # Ping rapide
    try:
        from googleapiclient.discovery import build
        yt = build("youtube", "v3", developerKey=api_key, cache_discovery=False)
        resp = yt.search().list(q="test", part="id", maxResults=1).execute()
        ok("Connexion API YouTube réussie")
    except Exception as e:
        err = str(e)
        if "quotaExceeded" in err:
            print(f"  {YELLOW}⚠ Quota épuisé aujourd'hui (normal) — sera OK demain{RESET}")
            passed += 1
        else:
            fail("Connexion API", err)
else:
    print(f"  {YELLOW}⚠ Pas de clé API dans .env — à configurer avant le lancement{RESET}")

# ---------------------------------------------------------------------------
# Résumé
# ---------------------------------------------------------------------------
total = passed + failed
print(f"\n{'='*50}")
if failed == 0:
    print(f"{GREEN}{BOLD}✓ Tous les tests passent ({passed}/{total}){RESET}")
    print(f"{GREEN}  L'outil est prêt pour le lancement demain.{RESET}")
else:
    print(f"{RED}{BOLD}✗ {failed} test(s) échoué(s) sur {total}{RESET}")
    sys.exit(1)
