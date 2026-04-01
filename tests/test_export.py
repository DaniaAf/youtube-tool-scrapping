import io
import json

import openpyxl
import pytest

from youtube_scraper import COLUMNS, calculate_tier, compute_scores, export_csv, export_excel, export_json


@pytest.fixture
def mock_profiles():
    profiles = []
    for i in range(10):
        followers = (i + 1) * 15_000
        eng = 0.01 + i * 0.008
        mentions = i % 5
        ppw = 0.5 + i * 0.3
        trend = float(i * 10 - 20)  # -20, -10, 0, 10, 20, ...
        shorts_ratio = round(i / ((i + 1) * 3), 3) if (i + 1) * 3 > 0 else 0.0
        aq_ratio = 0.05 if i % 2 == 0 else 0.01
        se, sc, sp, sr, saq, ssc, sg = compute_scores(
            eng, mentions, ppw, views_trend_pct=trend, has_video_stats=True, has_views_trend=True,
            audience_quality_ratio=aq_ratio, shorts_ratio=shorts_ratio,
        )
        profiles.append(
            {
                "platform": "YouTube",
                "username": f"creator_{i}",
                "display_name": f"Test Creator {i}",
                "profile_url": f"https://www.youtube.com/@creator_{i}",
                "email": f"creator{i}@example.com",
                "bio_snippet": f"Bio of creator {i}.",
                "followers": followers,
                "tier": calculate_tier(followers),
                "engagement_rate": round(eng, 6),
                "engagement_rate_pct": round(eng * 100, 3),
                "views_trend_pct": trend,
                "posts_per_week": round(ppw, 2),
                "keyword_mentions": mentions,
                "is_emerging": trend > 15 and followers < 100_000,
                "score_global": sg,
                "score_engagement": se,
                "score_croissance": sc,
                "score_pertinence": sp,
                "score_regularite": sr,
                "score_audience_quality": saq,
                "score_shorts_content": ssc,
                "total_recent_views": (i + 1) * 1_000,
                "total_recent_likes": (i + 1) * 50,
                "total_recent_comments": (i + 1) * 5,
                "recent_video_count": (i + 1) * 3,
                "shorts_count": i,
                "long_form_count": (i + 1) * 3 - i,
                "shorts_ratio": shorts_ratio,
                "punch_above_weight": "normal",
                "punch_above_weight_ratio": 1.0,
                "content_categories": "Gaming, Sports" if i % 2 == 0 else "",
                "channel_keywords": f"keyword{i}" if i % 3 == 0 else "",
                "audience_quality": "good",
                "status": "active" if ppw >= 0.5 else "inactive",
                "collected_at": "2025-01-01 12:00:00",
            }
        )
    return profiles


class TestExportExcel:
    def test_export_to_buffer(self, mock_profiles):
        buf = io.BytesIO()
        export_excel(mock_profiles, buf, ["test"])
        buf.seek(0)
        assert len(buf.read()) > 1000

    def test_export_to_file(self, tmp_path, mock_profiles):
        filepath = tmp_path / "test_output.xlsx"
        export_excel(mock_profiles, str(filepath), ["test"])
        assert filepath.exists()
        assert filepath.stat().st_size > 1000

    def test_sheet_names(self, tmp_path, mock_profiles):
        filepath = tmp_path / "test_output.xlsx"
        export_excel(mock_profiles, str(filepath), ["test"])
        wb = openpyxl.load_workbook(filepath)
        assert "Profiles" in wb.sheetnames
        assert "Summary" in wb.sheetnames

    def test_column_headers(self, tmp_path, mock_profiles):
        filepath = tmp_path / "test_output.xlsx"
        export_excel(mock_profiles, str(filepath), ["test"])
        wb = openpyxl.load_workbook(filepath)
        ws = wb["Profiles"]
        headers = [ws.cell(1, c).value for c in range(1, len(COLUMNS) + 1)]
        assert headers == COLUMNS

    def test_row_count(self, tmp_path, mock_profiles):
        filepath = tmp_path / "test_output.xlsx"
        export_excel(mock_profiles, str(filepath), ["test"])
        wb = openpyxl.load_workbook(filepath)
        ws = wb["Profiles"]
        data_rows = ws.max_row - 1  # minus header
        assert data_rows == len(mock_profiles)


class TestExportCSV:
    def test_export_to_buffer(self, mock_profiles):
        buf = io.StringIO()
        export_csv(mock_profiles, buf, ["test"])
        content = buf.getvalue()
        lines = content.strip().split("\n")
        assert len(lines) == len(mock_profiles) + 1  # header + data

    def test_export_to_file(self, tmp_path, mock_profiles):
        filepath = tmp_path / "test_output.csv"
        export_csv(mock_profiles, str(filepath), ["test"])
        assert filepath.exists()
        lines = filepath.read_text().strip().split("\n")
        assert len(lines) == len(mock_profiles) + 1

    def test_header_matches_columns(self, tmp_path, mock_profiles):
        filepath = tmp_path / "test_output.csv"
        export_csv(mock_profiles, str(filepath), ["test"])
        header = filepath.read_text().split("\n")[0].strip()
        assert header == ",".join(COLUMNS)


class TestExportJSON:
    def test_export_to_buffer(self, mock_profiles):
        buf = io.BytesIO()
        export_json(mock_profiles, buf, ["test"])
        buf.seek(0)
        data = json.loads(buf.read().decode("utf-8"))
        assert "metadata" in data
        assert "profiles" in data
        assert data["metadata"]["total_profiles"] == len(mock_profiles)
        assert data["metadata"]["keywords"] == ["test"]

    def test_export_to_file(self, tmp_path, mock_profiles):
        filepath = tmp_path / "test_output.json"
        export_json(mock_profiles, str(filepath), ["test"])
        assert filepath.exists()
        data = json.loads(filepath.read_text())
        assert len(data["profiles"]) == len(mock_profiles)

    def test_profiles_sorted_by_score(self, tmp_path, mock_profiles):
        filepath = tmp_path / "test_output.json"
        export_json(mock_profiles, str(filepath), ["test"])
        data = json.loads(filepath.read_text())
        scores = [p["score_global"] for p in data["profiles"]]
        assert scores == sorted(scores, reverse=True)
