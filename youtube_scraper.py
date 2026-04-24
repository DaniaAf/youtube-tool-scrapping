#!/usr/bin/env python3
"""
YouTube Profile Scraper
========================
Fetches YouTube channel profiles based on keywords, country, and time period.
Outputs an Excel file with scoring indices.

Usage:
    python youtube_scraper.py --keywords "Sorare" --region FR --days 90
    python youtube_scraper.py --keywords "Sorare" "NFT" --region FR --days 90 --output results.xlsx
    python youtube_scraper.py --keywords "Sorare" --region FR --days 90 --max-channels 200
"""

import argparse
import io
import json as json_module
import logging
import os
import re
import sys
import time
from datetime import UTC, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import diskcache
import pandas as pd
from dotenv import load_dotenv
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tenacity import retry, retry_if_exception, stop_after_attempt, wait_exponential
from tqdm import tqdm

load_dotenv()

logger = logging.getLogger(__name__)

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
ISO8601_DURATION_RE = re.compile(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?")
SHORTS_MAX_DURATION_SECONDS = 60


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

TIER_BOUNDARIES = [(1_000_000, "mega"), (100_000, "macro"), (10_000, "mid"), (1_000, "micro"), (0, "nano")]
SCORE_WEIGHTS = {
    "pertinence": 0.25,
    "engagement": 0.22,
    "croissance": 0.18,
    "regularite": 0.12,
    "audience_quality": 0.13,
    "shorts_content": 0.10,
}
ENGAGEMENT_THRESHOLDS = [(0.10, 100), (0.07, 85), (0.05, 70), (0.03, 55), (0.01, 35), (0, 15)]
ENGAGEMENT_THRESHOLDS_BY_TIER = {
    "nano": [(0.15, 100), (0.10, 85), (0.07, 70), (0.05, 55), (0.02, 35), (0, 15)],
    "micro": [(0.12, 100), (0.08, 85), (0.06, 70), (0.04, 55), (0.015, 35), (0, 15)],
    "mid": ENGAGEMENT_THRESHOLDS,  # default
    "macro": [(0.06, 100), (0.04, 85), (0.03, 70), (0.02, 55), (0.008, 35), (0, 15)],
    "mega": [(0.04, 100), (0.03, 85), (0.02, 70), (0.01, 55), (0.005, 35), (0, 15)],
}
PERTINENCE_THRESHOLDS = [(10, 100), (5, 85), (3, 65), (2, 45), (1, 25), (0, 0)]
REGULARITE_THRESHOLDS = [(4, 100), (3, 85), (2, 70), (1, 50), (0.5, 30), (0, 10)]
VIEWS_TREND_THRESHOLDS = [(100, 100), (50, 85), (20, 70), (5, 55), (0, 40), (-20, 25), (-50, 15)]
EMERGING_TREND_MIN_PCT = 15
EMERGING_FOLLOWERS_MAX = 100_000
EMERGING_SCORE_BONUS = 8
EMERGING_PAW_MIN_RATIO = 3.0
EMERGING_MENTIONS_MIN = 1
EMERGING_PPW_MIN = 0.3
EMERGING_FALLBACK_MENTIONS_MIN = 2
EMERGING_FALLBACK_PPW_MIN = 0.5
RATE_LIMIT_SEARCH = 0.2
RATE_LIMIT_CHANNEL_DETAILS = 0.1
RATE_LIMIT_VIDEO_STATS = 0.15
ACTIVE_THRESHOLD_PPW = 0.5
AUDIENCE_QUALITY_THRESHOLDS = [(0.10, "excellent"), (0.03, "good"), (0.005, "average"), (0, "low")]
PAW_THRESHOLDS = [
    (10.0, "exceptional"), (3.0, "strong"), (1.0, "normal"), (0.3, "below"), (0, "weak"),
]
PAW_SCORE_THRESHOLDS = [
    (10.0, 100), (5.0, 85), (3.0, 70), (1.5, 55), (1.0, 40), (0.5, 25), (0, 10),
]
AUDIENCE_QUALITY_SCORE_THRESHOLDS = [
    (0.10, 100), (0.05, 85), (0.03, 70), (0.01, 50), (0.005, 35), (0.001, 20), (0, 10),
]
SHORTS_CONTENT_THRESHOLDS = [
    (0.9, 100), (0.7, 90), (0.5, 70), (0.3, 50), (0.1, 30), (0.0, 15),
]
ZERO_VIDEO_STATS = {
    "views": 0, "likes": 0, "comments": 0, "video_count": 0,
    "shorts_count": 0, "long_form_count": 0, "per_video_views": [],
    "is_chronological": False,
}
FAST_MODE_BATCH_SIZE = 50

# Geo signals per region: (languages, city/country keywords, flag emojis)
REGION_GEO_SIGNALS: dict[str, dict] = {
    "FR": {
        "languages": {"fr"},
        "keywords": {"france", "français", "française", "paris", "lyon", "marseille", "bordeaux", "toulouse", "fr", "hexagone"},
        "flags": {"🇫🇷"},
    },
    "BE": {
        "languages": {"fr", "nl", "de"},
        "keywords": {"belgique", "belgië", "belgium", "bruxelles", "brussels", "liège", "gent", "antwerp"},
        "flags": {"🇧🇪"},
    },
    "CH": {
        "languages": {"fr", "de", "it"},
        "keywords": {"suisse", "schweiz", "switzerland", "zürich", "genève", "geneva", "basel", "bern"},
        "flags": {"🇨🇭"},
    },
    "GB": {
        "languages": {"en"},
        "keywords": {"uk", "united kingdom", "england", "london", "manchester", "british", "scotland", "wales"},
        "flags": {"🇬🇧"},
    },
    "US": {
        "languages": {"en"},
        "keywords": {"usa", "united states", "america", "new york", "los angeles", "chicago", "houston", "american", "us-based"},
        "flags": {"🇺🇸"},
    },
    "DE": {
        "languages": {"de"},
        "keywords": {"deutschland", "germany", "german", "berlin", "münchen", "munich", "hamburg", "köln", "deutsch"},
        "flags": {"🇩🇪"},
    },
    "ES": {
        "languages": {"es"},
        "keywords": {"españa", "spain", "spanish", "madrid", "barcelona", "sevilla", "español", "española"},
        "flags": {"🇪🇸"},
    },
    "IT": {
        "languages": {"it"},
        "keywords": {"italia", "italy", "italian", "roma", "milano", "napoli", "italiano", "italiana"},
        "flags": {"🇮🇹"},
    },
    "BR": {
        "languages": {"pt"},
        "keywords": {"brasil", "brazil", "brasileiro", "brasileira", "são paulo", "rio", "português"},
        "flags": {"🇧🇷"},
    },
    "CA": {
        "languages": {"en", "fr"},
        "keywords": {"canada", "canadian", "toronto", "montreal", "vancouver", "québec", "ottawa"},
        "flags": {"🇨🇦"},
    },
}

# Cache TTLs (seconds)
CACHE_TTL_SEARCH = 4 * 3600  # 4 hours
CACHE_TTL_CHANNEL_DETAILS = 24 * 3600  # 24 hours
CACHE_TTL_VIDEO_STATS = 4 * 3600  # 4 hours
_CACHE_DIR = Path.home() / ".youtube_scraper" / "cache"
_cache: diskcache.Cache | None = None

# Quota tracking
YOUTUBE_DAILY_QUOTA = 10_000
QUOTA_COST_SEARCH = 100  # search.list
QUOTA_COST_LIST = 1  # channels.list, videos.list


def get_cache() -> diskcache.Cache:
    """Get or create the disk cache instance."""
    global _cache
    if _cache is None:
        _cache = diskcache.Cache(str(_CACHE_DIR))
    return _cache


def clear_cache() -> None:
    """Clear all cached API responses."""
    cache = get_cache()
    cache.clear()
    logger.info("Cache cleared")


# ---------------------------------------------------------------------------
# Quota tracking (daily, resets at midnight Pacific time)
# ---------------------------------------------------------------------------

_PACIFIC = ZoneInfo("America/Los_Angeles")


def _quota_cache_key() -> str:
    """Return today's quota cache key using Pacific time (YouTube resets at midnight PT)."""
    today = datetime.now(_PACIFIC).strftime("%Y-%m-%d")
    return f"quota_used:{today}"


def _quota_ttl_seconds() -> int:
    """Return seconds until next midnight Pacific time."""
    now = datetime.now(_PACIFIC)
    midnight = now.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)
    return max(int((midnight - now).total_seconds()), 1)


def record_quota_usage(units: int) -> None:
    """Atomically increment today's quota counter in the disk cache."""
    cache = get_cache()
    key = _quota_cache_key()
    try:
        cache.incr(key, delta=units)
    except KeyError:
        cache.set(key, units, expire=_quota_ttl_seconds())


def get_quota_used() -> int:
    """Return today's consumed quota units."""
    cache = get_cache()
    return cache.get(_quota_cache_key(), default=0)


def reset_quota() -> None:
    """Delete today's quota counter."""
    cache = get_cache()
    cache.delete(_quota_cache_key())


RETRY_STATUS_CODES = {429, 500, 503}
MAX_RETRY_ATTEMPTS = 3


def _is_retryable_http_error(exc: BaseException) -> bool:
    """Return True for transient HTTP errors worth retrying."""
    if isinstance(exc, HttpError):
        return exc.resp.status in RETRY_STATUS_CODES
    return False


_api_retry = retry(
    retry=retry_if_exception(_is_retryable_http_error),
    stop=stop_after_attempt(MAX_RETRY_ATTEMPTS),
    wait=wait_exponential(multiplier=1, min=2, max=30),
    reraise=True,
)


@_api_retry
def _execute_api_request(request, quota_cost: int = 1):
    """Execute a YouTube API request with retry on transient errors."""
    result = request.execute()
    record_quota_usage(quota_cost)
    return result


# ---------------------------------------------------------------------------
# YouTube API client
# ---------------------------------------------------------------------------


def get_youtube_client(api_key: str):
    return build("youtube", "v3", developerKey=api_key, cache_discovery=False)


# ---------------------------------------------------------------------------
# Data collection
# ---------------------------------------------------------------------------


def resolve_channel_urls(youtube, urls: list[str]) -> list[str]:
    """
    Resolve a list of YouTube channel URLs/handles to channel IDs.
    Supports: @handle, /channel/UC..., /c/name, bare channel IDs.
    Returns a list of channel IDs (duplicates removed, order preserved).
    """
    channel_ids = []
    handles_to_resolve = []

    for url in urls:
        url = url.strip()
        if not url:
            continue
        # Already a channel ID
        if re.match(r"^UC[\w-]{22}$", url):
            channel_ids.append(url)
            continue
        # Extract from URL
        handle_match = re.search(r"youtube\.com/@([\w.-]+)", url)
        channel_id_match = re.search(r"youtube\.com/channel/(UC[\w-]{22})", url)
        c_match = re.search(r"youtube\.com/c/([\w.-]+)", url)

        if channel_id_match:
            channel_ids.append(channel_id_match.group(1))
        elif handle_match:
            handles_to_resolve.append("@" + handle_match.group(1))
        elif c_match:
            handles_to_resolve.append(c_match.group(1))
        else:
            # Try treating the raw string as a handle
            clean = url.lstrip("@")
            handles_to_resolve.append("@" + clean)

    # Resolve handles via API (forHandle param, 1 unit each)
    for handle in handles_to_resolve:
        try:
            resp = _execute_api_request(
                youtube.channels().list(forHandle=handle, part="id"),
                quota_cost=QUOTA_COST_LIST,
            )
            items = resp.get("items", [])
            if items:
                channel_ids.append(items[0]["id"])
            else:
                logger.warning("Could not resolve handle: %s", handle)
        except HttpError as e:
            logger.warning("API error resolving handle %s: %s", handle, e)

    # Deduplicate preserving order
    seen = set()
    result = []
    for cid in channel_ids:
        if cid not in seen:
            seen.add(cid)
            result.append(cid)
    return result


def search_videos_by_keyword(
    youtube,
    keyword: str,
    region_code: str | None,
    days: int,
    language: str | None = None,
    max_channels: int = 150,
    use_cache: bool = True,
) -> dict[str, dict]:
    """
    Search YouTube videos by keyword + region + recency.
    region_code=None means worldwide (no region filter).
    Returns a dict keyed by channel_id with mention counts and video ids.
    """
    cache_key = f"search:{keyword}:{region_code}:{days}:{language}:{max_channels}"
    if use_cache:
        cache = get_cache()
        cached = cache.get(cache_key)
        if cached is not None:
            logger.info("  Cache hit for search '%s'", keyword)
            return cached

    published_after = (datetime.now(UTC) - timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")

    channels: dict[str, dict] = {}
    next_page_token = None

    while True:
        try:
            params = dict(
                q=keyword,
                type="video",
                publishedAfter=published_after,
                part="snippet",
                maxResults=50,
                order="relevance",
            )
            if region_code:
                params["regionCode"] = region_code
            if language:
                params["relevanceLanguage"] = language
            if next_page_token:
                params["pageToken"] = next_page_token

            response = _execute_api_request(youtube.search().list(**params), quota_cost=QUOTA_COST_SEARCH)

        except HttpError:
            raise  # propagate to caller for proper UI error handling

        for item in response.get("items", []):
            channel_id = item["snippet"]["channelId"]
            title = item["snippet"].get("title", "")
            description = item["snippet"].get("description", "")
            video_id = item["id"].get("videoId", "")

            if channel_id not in channels:
                channels[channel_id] = {
                    "channel_id": channel_id,
                    "display_name": item["snippet"]["channelTitle"],
                    "video_ids": [],
                    "mentions_count": 0,
                }

            # Count keyword mentions in title or truncated description snippet
            kw_lower = keyword.lower()
            if kw_lower in title.lower() or kw_lower in description.lower():
                channels[channel_id]["mentions_count"] += 1

            if video_id:
                channels[channel_id]["video_ids"].append(video_id)

        next_page_token = response.get("nextPageToken")
        if not next_page_token or len(channels) >= max_channels:
            break

        time.sleep(RATE_LIMIT_SEARCH)

    # Fetch full video descriptions to catch keyword mentions cut off in snippets
    # videos.list costs 1 quota per batch of 50 — very cheap
    all_video_ids = [vid for ch in channels.values() for vid in ch["video_ids"]]
    video_to_channel = {vid: ch["channel_id"] for ch in channels.values() for vid in ch["video_ids"]}
    # Reset mention counts — we'll recount from full descriptions
    for ch in channels.values():
        ch["mentions_count"] = 0

    kw_lower = keyword.lower()
    for i in range(0, len(all_video_ids), 50):
        batch = all_video_ids[i : i + 50]
        try:
            resp = _execute_api_request(
                youtube.videos().list(id=",".join(batch), part="snippet"),
                quota_cost=QUOTA_COST_LIST,
            )
        except HttpError:
            logger.warning("Failed to fetch full descriptions for keyword mention recount")
            break

        for item in resp.get("items", []):
            video_id = item["id"]
            channel_id = video_to_channel.get(video_id)
            if not channel_id or channel_id not in channels:
                continue
            title = item["snippet"].get("title", "")
            full_description = item["snippet"].get("description", "")
            if kw_lower in title.lower() or kw_lower in full_description.lower():
                channels[channel_id]["mentions_count"] += 1

        time.sleep(RATE_LIMIT_CHANNEL_DETAILS)

    if use_cache:
        cache = get_cache()
        cache.set(cache_key, channels, expire=CACHE_TTL_SEARCH)

    return channels


def _parse_topic_categories(topic_details: dict) -> list[str]:
    """Extract human-readable topic labels from topicDetails.topicCategories URLs."""
    urls = topic_details.get("topicCategories", [])
    labels = []
    for url in urls:
        # URLs look like "https://en.wikipedia.org/wiki/Gaming"
        if "/wiki/" in url:
            label = url.rsplit("/wiki/", 1)[-1].replace("_", " ")
            labels.append(label)
    return labels


def _parse_iso8601_duration(duration: str) -> int:
    """Parse ISO 8601 duration (e.g. PT1H2M3S) to total seconds."""
    match = ISO8601_DURATION_RE.match(duration)
    if not match:
        return 0
    hours = int(match.group(1) or 0)
    minutes = int(match.group(2) or 0)
    seconds = int(match.group(3) or 0)
    return hours * 3600 + minutes * 60 + seconds


def get_channel_details(youtube, channel_ids: list[str], use_cache: bool = True) -> dict[str, dict]:
    """Fetch channel metadata and statistics in batches of 50."""
    result = {}
    ids_to_fetch = []
    cache = get_cache() if use_cache else None

    # Check cache for each channel
    if cache is not None:
        for cid in channel_ids:
            cached = cache.get(f"channel:{cid}")
            if cached is not None:
                result[cid] = cached
            else:
                ids_to_fetch.append(cid)
    else:
        ids_to_fetch = list(channel_ids)

    if ids_to_fetch:
        logger.info("  Fetching %d channels from API (%d from cache)", len(ids_to_fetch), len(result))

    for i in range(0, len(ids_to_fetch), 50):
        batch = ids_to_fetch[i : i + 50]
        try:
            response = _execute_api_request(
                youtube.channels().list(id=",".join(batch), part="snippet,statistics,topicDetails,brandingSettings"),
                quota_cost=QUOTA_COST_LIST,
            )
        except HttpError:
            raise  # propagate to caller

        for item in response.get("items", []):
            cid = item["id"]
            stats = item.get("statistics", {})
            snippet = item.get("snippet", {})
            topic_details = item.get("topicDetails", {})
            branding = item.get("brandingSettings", {})

            subscribers = stats.get("subscriberCount")
            full_description = snippet.get("description", "")
            email_match = EMAIL_RE.search(full_description)

            content_categories = _parse_topic_categories(topic_details)
            channel_keywords = branding.get("channel", {}).get("keywords", "")

            result[cid] = {
                "username": snippet.get("customUrl", "").lstrip("@"),
                "display_name": snippet.get("title", ""),
                "bio_snippet": full_description[:250].replace("\n", " "),
                "email": email_match.group(0) if email_match else "",
                "country": snippet.get("country", ""),
                "default_language": snippet.get("defaultLanguage", ""),
                "published_at": snippet.get("publishedAt", ""),
                "followers": int(subscribers) if subscribers else 0,
                "total_views": int(stats.get("viewCount", 0) or 0),
                "total_video_count": int(stats.get("videoCount", 0) or 0),
                "hidden_subscribers": stats.get("hiddenSubscriberCount", False),
                "content_categories": content_categories,
                "channel_keywords": channel_keywords,
            }
            if cache is not None:
                cache.set(f"channel:{cid}", result[cid], expire=CACHE_TTL_CHANNEL_DETAILS)

        time.sleep(RATE_LIMIT_CHANNEL_DETAILS)

    return result


def get_recent_video_stats(youtube, channel_id: str, days: int, use_cache: bool = True) -> dict:
    """Fetch aggregate stats (views, likes, comments) for recent videos."""
    cache_key = f"vstats:{channel_id}:{days}"
    if use_cache:
        cache = get_cache()
        cached = cache.get(cache_key)
        if cached is not None:
            return cached

    published_after = (datetime.now(UTC) - timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")

    try:
        search_resp = _execute_api_request(
            youtube.search().list(
                channelId=channel_id,
                type="video",
                publishedAfter=published_after,
                part="id",
                maxResults=50,
                order="date",
            ),
            quota_cost=QUOTA_COST_SEARCH,
        )
    except HttpError:
        logger.warning("Failed to search videos for channel %s", channel_id)
        return dict(ZERO_VIDEO_STATS)

    video_ids = [item["id"]["videoId"] for item in search_resp.get("items", []) if item["id"].get("videoId")]

    if not video_ids:
        return dict(ZERO_VIDEO_STATS)

    try:
        stats_resp = _execute_api_request(
            youtube.videos().list(id=",".join(video_ids), part="statistics,contentDetails"),
            quota_cost=QUOTA_COST_LIST,
        )
    except HttpError:
        logger.warning("Failed to fetch video stats for channel %s", channel_id)
        return {
            "views": 0,
            "likes": 0,
            "comments": 0,
            "video_count": len(video_ids),
            "shorts_count": 0,
            "long_form_count": 0,
            "per_video_views": [],
            "is_chronological": False,
        }

    total_views = total_likes = total_comments = 0
    shorts_count = long_form_count = 0
    views_by_id: dict[str, int] = {}
    for item in stats_resp.get("items", []):
        s = item.get("statistics", {})
        views = int(s.get("viewCount", 0) or 0)
        views_by_id[item["id"]] = views
        total_views += views
        total_likes += int(s.get("likeCount", 0) or 0)
        total_comments += int(s.get("commentCount", 0) or 0)
        duration_str = item.get("contentDetails", {}).get("duration", "")
        duration_secs = _parse_iso8601_duration(duration_str)
        if 0 < duration_secs <= SHORTS_MAX_DURATION_SECONDS:
            shorts_count += 1
        else:
            long_form_count += 1

    # Chronological order (oldest → newest) for sparkline display
    per_video_views = [views_by_id.get(vid, 0) for vid in reversed(video_ids)]

    result = {
        "views": total_views,
        "likes": total_likes,
        "comments": total_comments,
        "video_count": len(video_ids),
        "shorts_count": shorts_count,
        "long_form_count": long_form_count,
        "per_video_views": per_video_views,
        "is_chronological": True,
    }
    if use_cache:
        cache = get_cache()
        cache.set(cache_key, result, expire=CACHE_TTL_VIDEO_STATS)
    return result


def get_video_stats_batch(youtube, video_ids: list[str]) -> dict:
    """Fetch aggregate stats for a list of video IDs (from keyword search).

    Uses videos().list() directly — no per-channel search needed.
    Batches in groups of FAST_MODE_BATCH_SIZE (50). Costs 1 unit per batch
    instead of 100 units per channel search.
    """
    if not video_ids:
        return dict(ZERO_VIDEO_STATS)

    total_views = total_likes = total_comments = 0
    fetched_count = 0
    shorts_count = long_form_count = 0
    per_video_views: list[int] = []

    for i in range(0, len(video_ids), FAST_MODE_BATCH_SIZE):
        batch = video_ids[i : i + FAST_MODE_BATCH_SIZE]
        try:
            resp = _execute_api_request(
                youtube.videos().list(id=",".join(batch), part="statistics,contentDetails"),
                quota_cost=QUOTA_COST_LIST,
            )
        except HttpError:
            logger.warning("Failed to fetch video stats batch (offset %d)", i)
            continue

        for item in resp.get("items", []):
            s = item.get("statistics", {})
            views = int(s.get("viewCount", 0) or 0)
            per_video_views.append(views)
            total_views += views
            total_likes += int(s.get("likeCount", 0) or 0)
            total_comments += int(s.get("commentCount", 0) or 0)
            fetched_count += 1
            duration_str = item.get("contentDetails", {}).get("duration", "")
            duration_secs = _parse_iso8601_duration(duration_str)
            if 0 < duration_secs <= SHORTS_MAX_DURATION_SECONDS:
                shorts_count += 1
            else:
                long_form_count += 1

        time.sleep(RATE_LIMIT_VIDEO_STATS)

    return {
        "views": total_views,
        "likes": total_likes,
        "comments": total_comments,
        "video_count": fetched_count,
        "shorts_count": shorts_count,
        "long_form_count": long_form_count,
        "per_video_views": per_video_views,
        "is_chronological": False,
    }


# ---------------------------------------------------------------------------
# Views trend
# ---------------------------------------------------------------------------


def _compute_views_trend(per_video_views: list[int]) -> float | None:
    """Compute % change between avg views of older vs newer halves of videos.

    Videos must be in chronological order (oldest first).
    Returns None when insufficient data.
    """
    if len(per_video_views) < 2:
        return None
    mid = len(per_video_views) // 2
    older = per_video_views[:mid]
    newer = per_video_views[mid:]
    avg_older = sum(older) / len(older)
    avg_newer = sum(newer) / len(newer)
    if avg_older == 0 and avg_newer == 0:
        return None
    if avg_older == 0:
        return 200.0
    return (avg_newer - avg_older) / avg_older * 100


# ---------------------------------------------------------------------------
# Scoring helpers
# ---------------------------------------------------------------------------


def calculate_tier(followers: int) -> str:
    for boundary, tier in TIER_BOUNDARIES:
        if followers >= boundary:
            return tier
    return "nano"


def compute_local_confidence(details: dict, region_code: str | None) -> str:
    """
    Estimate how likely a creator is based in / targeting the given region.
    Returns: "high", "medium", "low", or "unknown" (when region is None/Worldwide).

    Scoring (additive, max 3 points → high):
    - +2 if declared country matches region
    - +1 if declared language matches region languages
    - +1 if geo keyword found in bio or channel_keywords
    - +1 if flag emoji found in bio or channel_keywords
    """
    if not region_code or region_code not in REGION_GEO_SIGNALS:
        return "unknown"

    signals = REGION_GEO_SIGNALS[region_code]
    score = 0

    # Country declared
    country = details.get("country", "").upper()
    if country and country == region_code.upper():
        score += 2

    # Language match
    lang = details.get("default_language", "").lower().split("-")[0]
    if lang and lang in signals["languages"]:
        score += 1

    # Keyword / flag scan in bio + channel keywords
    text = (details.get("bio_snippet", "") + " " + details.get("channel_keywords", "")).lower()
    if any(kw in text for kw in signals["keywords"]):
        score += 1
    if any(flag in text for flag in signals["flags"]):
        score += 1

    if score >= 2:
        return "high"
    elif score == 1:
        return "medium"
    else:
        return "low"


def classify_audience_quality(followers: int, total_views: int) -> str:
    """Classify audience quality based on subscriber-to-view ratio."""
    if total_views <= 0:
        return "unknown"
    ratio = followers / total_views
    for min_ratio, label in AUDIENCE_QUALITY_THRESHOLDS:
        if ratio >= min_ratio:
            return label
    return "low"


def compute_punch_above_weight(followers: int, per_video_views: list[int]) -> tuple[float, str]:
    """Compute PAW ratio (avg views per video / followers) and classify it."""
    if not per_video_views or followers <= 0:
        return (0.0, "unknown")
    avg_views = sum(per_video_views) / len(per_video_views)
    ratio = round(avg_views / followers, 3)
    label = "weak"
    for min_ratio, lbl in PAW_THRESHOLDS:
        if ratio >= min_ratio:
            label = lbl
            break
    return (ratio, label)


def _score_from_thresholds(value: float, thresholds: list[tuple[float, float]], default: float = 0) -> float:
    for min_value, score in thresholds:
        if value >= min_value:
            return score
    return default


def score_engagement(rate: float, tier: str | None = None) -> float:
    """Rate is a ratio (e.g. 0.05 = 5%). When tier is provided, uses tier-specific thresholds."""
    thresholds = ENGAGEMENT_THRESHOLDS_BY_TIER.get(tier, ENGAGEMENT_THRESHOLDS) if tier else ENGAGEMENT_THRESHOLDS
    return _score_from_thresholds(rate, thresholds, default=15)


def score_pertinence(mentions: int) -> float:
    return _score_from_thresholds(mentions, PERTINENCE_THRESHOLDS, default=0)


def score_regularite(posts_per_week: float) -> float:
    return _score_from_thresholds(posts_per_week, REGULARITE_THRESHOLDS, default=10)


def score_croissance(views_trend_pct: float) -> float:
    return _score_from_thresholds(views_trend_pct, VIEWS_TREND_THRESHOLDS, default=10)


def score_audience_quality(subscriber_view_ratio: float) -> float:
    """Score audience quality from subscriber/view ratio. Higher ratio = more loyal audience."""
    return _score_from_thresholds(subscriber_view_ratio, AUDIENCE_QUALITY_SCORE_THRESHOLDS, default=10)


def score_shorts_content(shorts_ratio: float) -> float:
    """Score content format mix. Penalizes high shorts ratio (brands prefer long-form)."""
    return _score_from_thresholds(1.0 - shorts_ratio, SHORTS_CONTENT_THRESHOLDS, default=15)


def compute_scores(
    engagement_rate,
    mentions,
    posts_per_week,
    views_trend_pct: float | None = None,
    has_video_stats=False,
    has_views_trend: bool = False,
    tier=None,
    audience_quality_ratio: float | None = None,
    shorts_ratio: float | None = None,
):
    w = SCORE_WEIGHTS
    sp = score_pertinence(mentions)
    sr = score_regularite(posts_per_week)

    # Build active weights based on available data
    active: dict[str, float] = {"pertinence": w["pertinence"], "regularite": w["regularite"]}
    if has_video_stats:
        active["engagement"] = w["engagement"]
    if has_views_trend and views_trend_pct is not None:
        active["croissance"] = w["croissance"]
    if audience_quality_ratio is not None:
        active["audience_quality"] = w["audience_quality"]
    if has_video_stats and shorts_ratio is not None:
        active["shorts_content"] = w["shorts_content"]

    total = sum(active.values())
    norm = {k: v / total for k, v in active.items()}

    se = score_engagement(engagement_rate, tier=tier) if "engagement" in norm else 0
    sc = score_croissance(views_trend_pct) if "croissance" in norm and views_trend_pct is not None else 0
    saq = score_audience_quality(audience_quality_ratio) if "audience_quality" in norm else 0
    ssc = score_shorts_content(shorts_ratio) if "shorts_content" in norm else 0

    sg = round(
        sp * norm["pertinence"]
        + se * norm.get("engagement", 0)
        + sc * norm.get("croissance", 0)
        + sr * norm["regularite"]
        + saq * norm.get("audience_quality", 0)
        + ssc * norm.get("shorts_content", 0),
        1,
    )

    return round(se, 1), round(sc, 1), round(sp, 1), round(sr, 1), round(saq, 1), round(ssc, 1), sg


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

COLUMNS = [
    "platform",
    "username",
    "display_name",
    "profile_url",
    "email",
    "bio_snippet",
    "creator_country",
    "target_market",
    "market_match",
    "local_confidence",
    "followers",
    "tier",
    "engagement_rate",
    "engagement_rate_pct",
    "views_trend_pct",
    "posts_per_week",
    "keyword_mentions",
    "is_emerging",
    "score_global",
    "score_engagement",
    "score_croissance",
    "score_pertinence",
    "score_regularite",
    "score_audience_quality",
    "score_shorts_content",
    "total_recent_views",
    "total_recent_likes",
    "total_recent_comments",
    "recent_video_count",
    "shorts_count",
    "long_form_count",
    "shorts_ratio",
    "punch_above_weight",
    "punch_above_weight_ratio",
    "content_categories",
    "channel_keywords",
    "audience_quality",
    "status",
    "collected_at",
]

SCORE_COLS = [
    "score_global", "score_engagement", "score_croissance", "score_pertinence",
    "score_regularite", "score_audience_quality", "score_shorts_content",
]

TIER_COLORS = {
    "mega": "7B2FBE",
    "macro": "3B82F6",
    "mid": "10B981",
    "micro": "F59E0B",
    "nano": "6B7280",
}

HEADER_BG = "1E3A5F"
ALT_ROW_BG = "EEF4FB"


def export_excel(profiles: list[dict], output_file, keywords: list[str]):
    df = pd.DataFrame(profiles, columns=COLUMNS)
    df.sort_values("score_global", ascending=False, inplace=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Profiles")

        wb = writer.book
        ws = writer.sheets["Profiles"]

        # ---- Header styling ----
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_fill = PatternFill("solid", fgColor=HEADER_BG)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws.row_dimensions[1].height = 32

        # ---- Column widths ----
        col_widths = {
            "platform": 12,
            "username": 22,
            "display_name": 28,
            "profile_url": 40,
            "email": 30,
            "bio_snippet": 45,
            "followers": 14,
            "tier": 10,
            "engagement_rate": 18,
            "engagement_rate_pct": 20,
            "views_trend_pct": 18,
            "posts_per_week": 16,
            "keyword_mentions": 18,
            "is_emerging": 14,
            "score_global": 14,
            "score_engagement": 18,
            "score_croissance": 18,
            "score_pertinence": 18,
            "score_regularite": 18,
            "score_audience_quality": 22,
            "score_shorts_content": 20,
            "total_recent_views": 18,
            "total_recent_likes": 18,
            "total_recent_comments": 20,
            "recent_video_count": 18,
            "shorts_count": 14,
            "long_form_count": 16,
            "shorts_ratio": 14,
            "punch_above_weight": 20,
            "punch_above_weight_ratio": 22,
            "content_categories": 30,
            "channel_keywords": 30,
            "audience_quality": 18,
            "status": 12,
            "collected_at": 20,
        }
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 16)

        # ---- Row styling ----
        col_map = {name: idx + 1 for idx, name in enumerate(COLUMNS)}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            # Alternating row background
            bg = ALT_ROW_BG if row_idx % 2 == 0 else "FFFFFF"
            alt_fill = PatternFill("solid", fgColor=bg)

            for cell in row:
                cell.fill = alt_fill
                cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Tier badge color
            tier_cell = ws.cell(row=row_idx, column=col_map["tier"])
            tier_val = tier_cell.value or ""
            color = TIER_COLORS.get(tier_val, "6B7280")
            tier_cell.font = Font(color=color, bold=True)

            # Hyperlink on display_name (nom cliquable → chaîne YouTube)
            name_cell = ws.cell(row=row_idx, column=col_map["display_name"])
            url_cell = ws.cell(row=row_idx, column=col_map["profile_url"])
            if url_cell.value:
                name_cell.hyperlink = url_cell.value
                name_cell.font = Font(color="1155CC", underline="single", bold=True)

            # Hyperlink on profile_url également
            if url_cell.value:
                url_cell.hyperlink = url_cell.value
                url_cell.font = Font(color="1155CC", underline="single")

            # is_emerging boolean → yes/no
            em_cell = ws.cell(row=row_idx, column=col_map["is_emerging"])
            em_cell.value = "Yes" if em_cell.value else "No"
            if em_cell.value == "Yes":
                em_cell.font = Font(color="10B981", bold=True)

        # ---- Color scale on score columns ----
        for score_col in SCORE_COLS:
            col_letter = get_column_letter(col_map[score_col])
            cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
            rule = ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="FF4444",
                mid_type="num",
                mid_value=50,
                mid_color="FFAA00",
                end_type="num",
                end_value=100,
                end_color="00CC44",
            )
            ws.conditional_formatting.add(cell_range, rule)

        # ---- Summary sheet ----
        ws_sum = wb.create_sheet("Summary")
        ws_sum["A1"] = "YouTube Scraper — Summary"
        ws_sum["A1"].font = Font(bold=True, size=14, color=HEADER_BG)

        summary_data = [
            ("Generated at", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Keywords", ", ".join(keywords)),
            ("Total profiles", len(profiles)),
            ("", ""),
            ("Tier breakdown", ""),
        ]
        tier_counts = df["tier"].value_counts().to_dict()
        for tier in ["mega", "macro", "mid", "micro", "nano"]:
            summary_data.append((f"  {tier}", tier_counts.get(tier, 0)))

        summary_data += [
            ("", ""),
            ("Avg score_global", round(df["score_global"].mean(), 1)),
            ("Avg engagement_rate_pct", round(df["engagement_rate_pct"].mean(), 2)),
            ("Channels with mentions > 0", int((df["keyword_mentions"] > 0).sum())),
            ("Emerging channels", int((df["is_emerging"] == "Yes").sum())),
        ]

        for r, (label, value) in enumerate(summary_data, start=3):
            ws_sum.cell(row=r, column=1, value=label).font = Font(bold=bool(label and not label.startswith(" ")))
            ws_sum.cell(row=r, column=2, value=value)

        ws_sum.column_dimensions["A"].width = 30
        ws_sum.column_dimensions["B"].width = 25

    logger.info("Saved → %s", output_file)


def export_csv(profiles: list[dict], output_file, keywords: list[str]):
    """Export profiles to CSV format."""
    df = pd.DataFrame(profiles, columns=COLUMNS)
    df.sort_values("score_global", ascending=False, inplace=True)
    if hasattr(output_file, "write"):
        df.to_csv(output_file, index=False)
    else:
        df.to_csv(output_file, index=False)
    logger.info("Saved → %s", output_file)


def export_json(profiles: list[dict], output_file, keywords: list[str]):
    """Export profiles to JSON format with metadata wrapper."""
    df = pd.DataFrame(profiles, columns=COLUMNS)
    df.sort_values("score_global", ascending=False, inplace=True)

    output = {
        "metadata": {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "keywords": keywords,
            "total_profiles": len(profiles),
        },
        "profiles": df.to_dict(orient="records"),
    }

    json_str = json_module.dumps(output, indent=2, default=str)
    if hasattr(output_file, "write"):
        output_file.write(json_str.encode("utf-8") if isinstance(output_file, io.BytesIO) else json_str)
    else:
        with open(output_file, "w") as f:
            f.write(json_str)
    logger.info("Saved → %s", output_file)


# ---------------------------------------------------------------------------
# Shared logic (used by both scrape() and app.py)
# ---------------------------------------------------------------------------


def merge_keyword_results(all_channels: dict[str, dict], new_channels: dict[str, dict]) -> None:
    """Merge new keyword search results into all_channels, deduplicating by channel id and video id."""
    for cid, data in new_channels.items():
        if cid not in all_channels:
            all_channels[cid] = data
        else:
            all_channels[cid]["mentions_count"] += data["mentions_count"]
            existing_ids = set(all_channels[cid]["video_ids"])
            for vid in data["video_ids"]:
                if vid not in existing_ids:
                    all_channels[cid]["video_ids"].append(vid)
                    existing_ids.add(vid)


def compute_channel_metrics(details: dict, vstats: dict, search_data: dict, days: int) -> dict:
    """Compute engagement rate, posts/week, views trend, emerging flag from raw data."""
    followers = details.get("followers", 0)
    total_views = vstats["views"]
    total_likes = vstats["likes"]
    total_comments = vstats["comments"]
    video_count = vstats["video_count"]

    engagement_rate = (total_likes + total_comments) / total_views if total_views > 0 else 0.0
    weeks = days / 7
    posts_per_week = round(video_count / weeks, 2) if weeks > 0 else 0

    # Views trend: only meaningful with chronological video data
    is_chronological = vstats.get("is_chronological", False)
    per_video_views = vstats.get("per_video_views", [])
    if is_chronological and len(per_video_views) >= 2:
        views_trend_pct = _compute_views_trend(per_video_views)
    else:
        views_trend_pct = None
    has_views_trend = views_trend_pct is not None

    mentions = search_data.get("mentions_count", 0)

    shorts = vstats.get("shorts_count", 0)
    long_form = vstats.get("long_form_count", 0)
    total_classified = shorts + long_form
    shorts_ratio = round(shorts / total_classified, 3) if total_classified > 0 else 0.0

    total_all_time_views = details.get("total_views", 0)
    audience_quality = classify_audience_quality(followers, total_all_time_views)
    audience_quality_ratio = followers / total_all_time_views if total_all_time_views > 0 else None

    paw_ratio, paw_label = compute_punch_above_weight(followers, per_video_views)

    # Emerging: three paths (any one triggers)
    is_emerging = False
    if followers < EMERGING_FOLLOWERS_MAX:
        # Growth path (full mode): strong views trend + some relevance
        growth_path = (
            has_views_trend
            and views_trend_pct >= EMERGING_TREND_MIN_PCT
            and mentions >= EMERGING_MENTIONS_MIN
            and posts_per_week >= EMERGING_PPW_MIN
        )
        # Fallback path (no trend): higher relevance bar
        fallback_path = (
            not has_views_trend
            and mentions >= EMERGING_FALLBACK_MENTIONS_MIN
            and posts_per_week >= EMERGING_FALLBACK_PPW_MIN
        )
        # PAW path: high punch-above-weight ratio + some relevance
        paw_path = (
            paw_ratio >= EMERGING_PAW_MIN_RATIO
            and mentions >= EMERGING_MENTIONS_MIN
            and posts_per_week >= EMERGING_PPW_MIN
        )
        is_emerging = growth_path or fallback_path or paw_path

    return {
        "engagement_rate": engagement_rate,
        "posts_per_week": posts_per_week,
        "views_trend_pct": views_trend_pct,
        "has_views_trend": has_views_trend,
        "mentions": mentions,
        "is_emerging": is_emerging,
        "followers": followers,
        "total_recent_views": total_views,
        "total_recent_likes": total_likes,
        "total_recent_comments": total_comments,
        "recent_video_count": video_count,
        "shorts_count": shorts,
        "long_form_count": long_form,
        "shorts_ratio": shorts_ratio,
        "audience_quality": audience_quality,
        "audience_quality_ratio": audience_quality_ratio,
        "punch_above_weight": paw_label,
        "punch_above_weight_ratio": paw_ratio,
        "per_video_views": per_video_views,
    }


def build_channel_profile(
    cid: str,
    details: dict,
    search_data: dict,
    metrics: dict,
    has_video_stats: bool,
    collected_at: str,
    region_code: str | None = None,
) -> dict:
    """Build a full profile dict with correct scoring, status, and email field."""
    m = metrics
    tier = calculate_tier(m["followers"])
    se, sc, sp, sr, saq, ssc, sg = compute_scores(
        m["engagement_rate"],
        m["mentions"],
        m["posts_per_week"],
        views_trend_pct=m["views_trend_pct"],
        has_video_stats=has_video_stats,
        has_views_trend=m["has_views_trend"],
        tier=tier,
        audience_quality_ratio=m.get("audience_quality_ratio"),
        shorts_ratio=m["shorts_ratio"] if has_video_stats else None,
    )

    # Emerging bonus
    if m["is_emerging"]:
        sg = min(sg + EMERGING_SCORE_BONUS, 100)

    username = details.get("username") or cid
    display_name = details.get("display_name") or search_data.get("display_name", "")

    if details.get("username"):
        profile_url = f"https://www.youtube.com/@{details['username'].lstrip('@')}"
    else:
        profile_url = f"https://www.youtube.com/channel/{cid}"

    # Status: when video stats are not fetched, we can't know posting frequency — default to "active"
    if has_video_stats:
        status = "active" if m["posts_per_week"] >= ACTIVE_THRESHOLD_PPW else "inactive"
    else:
        status = "active"

    creator_country = details.get("country", "")
    target_market = region_code or "Worldwide"
    market_match = (
        creator_country.upper() == target_market.upper()
        if creator_country and target_market != "Worldwide"
        else None
    )
    local_confidence = compute_local_confidence(details, region_code)

    return {
        "platform": "YouTube",
        "username": username,
        "display_name": display_name,
        "profile_url": profile_url,
        "email": details.get("email", ""),
        "bio_snippet": details.get("bio_snippet", ""),
        "creator_country": creator_country,
        "target_market": target_market,
        "market_match": market_match,
        "local_confidence": local_confidence,
        "followers": m["followers"],
        "tier": calculate_tier(m["followers"]),
        "engagement_rate": round(m["engagement_rate"], 6),
        "engagement_rate_pct": round(m["engagement_rate"] * 100, 3),
        "views_trend_pct": m["views_trend_pct"],
        "posts_per_week": m["posts_per_week"],
        "keyword_mentions": m["mentions"],
        "is_emerging": m["is_emerging"],
        "score_global": sg,
        "score_engagement": se,
        "score_croissance": sc,
        "score_pertinence": sp,
        "score_regularite": sr,
        "score_audience_quality": saq,
        "score_shorts_content": ssc,
        "total_recent_views": m["total_recent_views"],
        "total_recent_likes": m["total_recent_likes"],
        "total_recent_comments": m["total_recent_comments"],
        "recent_video_count": m["recent_video_count"],
        "shorts_count": m["shorts_count"],
        "long_form_count": m["long_form_count"],
        "shorts_ratio": m["shorts_ratio"],
        "punch_above_weight": m["punch_above_weight"],
        "punch_above_weight_ratio": m["punch_above_weight_ratio"],
        "content_categories": ", ".join(details.get("content_categories", [])),
        "channel_keywords": details.get("channel_keywords", ""),
        "audience_quality": m["audience_quality"],
        "per_video_views": m["per_video_views"],
        "status": status,
        "collected_at": collected_at,
    }


# ---------------------------------------------------------------------------
# Main orchestration
# ---------------------------------------------------------------------------


def scrape(
    keywords: list[str],
    region_code: str = "FR",
    days: int = 90,
    language: str | None = None,
    api_key: str | None = None,
    output_file: str = "youtube_profiles.xlsx",
    max_channels: int = 150,
    fetch_video_stats: bool = True,
    video_stats_mode: str = "full",
    use_cache: bool = True,
    export_format: str = "xlsx",
):
    api_key = api_key or os.environ.get("YOUTUBE_API_KEY")
    if not api_key:
        sys.exit("ERROR: Set YOUTUBE_API_KEY env var or pass --api-key")

    youtube = get_youtube_client(api_key)

    # 1. Collect channels from keyword searches
    all_channels: dict[str, dict] = {}

    for kw in keywords:
        logger.info("[1/3] Searching '%s' | region=%s | last %dd …", kw, region_code, days)
        found = search_videos_by_keyword(youtube, kw, region_code, days, language, max_channels, use_cache=use_cache)
        logger.info("      → %d channels found", len(found))
        merge_keyword_results(all_channels, found)

    if not all_channels:
        logger.warning("No channels found. Try different keywords or a wider date range.")
        return

    channel_ids = list(all_channels.keys())[:max_channels]
    logger.info("[2/3] Fetching details for %d channels …", len(channel_ids))
    channel_details = get_channel_details(youtube, channel_ids, use_cache=use_cache)

    # 2. Build profiles
    logger.info("[3/3] Computing metrics …")
    profiles = []
    collected_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Resolve effective mode: legacy fetch_video_stats=False maps to "none"
    effective_mode = video_stats_mode
    if not fetch_video_stats and video_stats_mode == "full":
        effective_mode = "none"

    for cid in tqdm(channel_ids, unit="channel"):
        details = channel_details.get(cid, {})
        search_data = all_channels[cid]

        if effective_mode == "full":
            try:
                vstats = get_recent_video_stats(youtube, cid, days, use_cache=use_cache)
                time.sleep(RATE_LIMIT_VIDEO_STATS)
            except HttpError:
                vstats = dict(ZERO_VIDEO_STATS)
        elif effective_mode == "fast":
            vstats = get_video_stats_batch(youtube, search_data.get("video_ids", []))
        else:  # "none"
            vstats = dict(ZERO_VIDEO_STATS)

        has_stats = effective_mode != "none"
        metrics = compute_channel_metrics(details, vstats, search_data, days)
        profile = build_channel_profile(cid, details, search_data, metrics, has_stats, collected_at, region_code)
        profiles.append(profile)

    logger.info("%d profiles collected.", len(profiles))
    exporters = {"xlsx": export_excel, "csv": export_csv, "json": export_json}
    exporter = exporters.get(export_format, export_excel)
    exporter(profiles, output_file, keywords)
    return profiles


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    logging.basicConfig(level=logging.INFO, format="%(message)s")

    parser = argparse.ArgumentParser(description="Scrape YouTube profiles by keyword, country, and time period.")
    parser.add_argument(
        "--keywords",
        "-k",
        nargs="+",
        required=True,
        help='One or more keywords, e.g. --keywords "Sorare" "NFT foot"',
    )
    parser.add_argument(
        "--region",
        "-r",
        default="FR",
        help="ISO 3166-1 alpha-2 country code (default: FR)",
    )
    parser.add_argument(
        "--days",
        "-d",
        type=int,
        default=90,
        help="Look-back window in days (default: 90)",
    )
    parser.add_argument(
        "--language",
        "-l",
        default=None,
        help='Relevance language hint, e.g. "fr" (optional)',
    )
    parser.add_argument(
        "--output",
        "-o",
        default="youtube_profiles.xlsx",
        help="Output Excel file name (default: youtube_profiles.xlsx)",
    )
    parser.add_argument(
        "--max-channels",
        "-m",
        type=int,
        default=150,
        help="Max channels to retrieve per keyword (default: 150)",
    )
    parser.add_argument(
        "--api-key",
        default=None,
        help="YouTube Data API v3 key (overrides YOUTUBE_API_KEY env var)",
    )
    parser.add_argument(
        "--no-video-stats",
        action="store_true",
        help="Skip per-channel video stats fetch (backward compat, same as --video-stats-mode none)",
    )
    parser.add_argument(
        "--video-stats-mode",
        choices=["none", "fast", "full"],
        default="full",
        help="Video stats mode: none (skip), fast (reuse search video IDs), full (per-channel search). Default: full",
    )
    parser.add_argument(
        "--no-cache",
        action="store_true",
        help="Disable API response caching (forces fresh API calls)",
    )
    parser.add_argument(
        "--format",
        "-f",
        choices=["xlsx", "csv", "json"],
        default="xlsx",
        dest="export_format",
        help="Output format (default: xlsx)",
    )

    args = parser.parse_args()

    # --no-video-stats overrides --video-stats-mode for backward compat
    mode = "none" if args.no_video_stats else args.video_stats_mode

    scrape(
        keywords=args.keywords,
        region_code=args.region.upper(),
        days=args.days,
        language=args.language,
        api_key=args.api_key,
        output_file=args.output,
        max_channels=args.max_channels,
        fetch_video_stats=not args.no_video_stats,
        video_stats_mode=mode,
        use_cache=not args.no_cache,
        export_format=args.export_format,
    )


if __name__ == "__main__":
    main()
